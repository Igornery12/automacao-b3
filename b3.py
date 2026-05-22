from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import sqlite3
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
from openpyxl import Workbook, load_workbook
import os





"""

Before running the code, you need to enter the sender and recipient for sending the action report.

"""


conn = sqlite3.connect('b3_data/data.db')
cursor = conn.cursor()
cursor.execute("""
    CREATE TABLE IF NOT EXISTS investiment(
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    simbolo TEXT,
    preco TEXT,
    oscilação TEXT,
    data TEXT,
    hora TEXT)
""")

conn.commit()

driver = webdriver.Chrome()
driver.maximize_window()
driver.get('https://www.b3.com.br/pt_br/para-voce')

driver.find_element(By.CSS_SELECTOR,'a.btn-primary').click()
WebDriverWait(driver,15).until(
    EC.presence_of_element_located(
        (By.XPATH,'//*[@id="conteudo-principal"]/div[3]/div/div/div[2]/div/dl/li[4]/a'))
)

driver.find_element(By.XPATH,'//*[@id="conteudo-principal"]/div[3]/div/div/div[2]/div/dl/li[4]/a').click()

arquivo_excel = "relatorio.xlsx"

def criar_excel():

    if not os.path.exists(arquivo_excel):

        wb = Workbook()
        ws = wb.active
        ws.title = "Relatorio B3"

        ws.append([
            "Ativo",
            "Preço",
            "Oscilação",
            "Data",
            "Hora",
            "Status"
        ])

        wb.save(arquivo_excel)

def adicionar_excel(ativo, preco, oscilacao, data, hora, status):

    wb = load_workbook(arquivo_excel)
    ws = wb.active

    ws.append([
        ativo,
        preco,
        oscilacao,
        data,
        hora,
        status
    ])

    wb.save(arquivo_excel)

criar_excel()

def pesquisa(p):
    time.sleep(6)
    
    campo = driver.find_element(By.XPATH,'//*[@id="txtCampoPesquisa"]')
    campo.clear()   
    campo.send_keys(p)

    try:
        WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "onetrust-accept-btn-handler"))
        ).click()
    except:
        pass

    driver.find_element(By.XPATH,'//*[@id="btnBuscarOutrosAtivos"]').click()
    time.sleep(6)
    driver.find_element(By.XPATH,'//*[@id="btnBuscarOutrosAtivos"]').click()
    time.sleep(6)


    name = driver.find_element(By.XPATH,'//*[@id="ativo"]').text
    price = driver.find_element(By.XPATH,'//*[@id="cotacaoAtivo"]').text
    oscilation = driver.find_element(By.XPATH,'//*[@id="oscilacaoAtivo"]').text
    day = driver.find_element(By.XPATH,'//*[@id="dataConsulta"]').text
    hour = driver.find_element(By.XPATH,'//*[@id="horaConsulta"]').text
    if name != '______':
        cursor.execute("""
                INSERT INTO investiment (simbolo, preco,oscilação, data, hora)
                VALUES (?, ?, ?,?, ?)
            """, (name, f'R${price}',f'{oscilation}%', day, hour))
        
        adicionar_excel(
            name,
           f'R$ {price}',
            oscilation,
            day,
            hour,
            "Encontrado"
        )

    else:

        
        print( f'{p} não encontrado ou não existe ')

        adicionar_excel(
            p,
            '----',
            '----',
            '----',
            '----',
            "Não encontrado"
        )
    
    time.sleep(4)

    conn.commit()

    print(name)
    print(price)
    print(oscilation)
    print(day)
    print(hour)




def enviar_email(caminho_arquivo):

    remetente = ""
    senha = ""
    destinatario = ""

    msg = MIMEMultipart()

    msg["From"] = remetente
    msg["To"] = destinatario
    msg["Subject"] = "Relatório B3"

    corpo = "Segue o relatório Excel em anexo."
    msg.attach(MIMEText(corpo, "plain"))

    
    with open(caminho_arquivo, "rb") as arquivo:
        parte = MIMEBase("application", "octet-stream")
        parte.set_payload(arquivo.read())

    encoders.encode_base64(parte)

    parte.add_header(
        "Content-Disposition",
        f'attachment; filename="{caminho_arquivo}"'
    )

    msg.attach(parte)

    
    with smtplib.SMTP("smtp.gmail.com", 587) as servidor:
        servidor.starttls()
        servidor.login(remetente, senha)
        servidor.send_message(msg)

    print("Email enviado com sucesso!")



lista = ['abev3','petr4','petr5']
for p in lista:
    produto = pesquisa(p)

enviar_email("relatório.xlsx")

conn.close()
driver.close()